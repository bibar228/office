from docxtpl import DocxTemplate
import openpyxl
from time import gmtime, strftime

months = {"01": "января", "02": "февраля", "03": "марта", "04": "апреля", "05": "мая", "06": "июня",
              "07": "июля", "08": "августа", "09": "сентября", "10": "октября", "11": "ноября", "12": "декабря"}


def create_files():
    wookbook = openpyxl.load_workbook("Список заявок СК Репка.xlsx")
    ws = wookbook.active

    for row in range(2, ws.max_row):
        print("ФИО1", ws.cell(row, 3).value)
        print("ФИОм", f"{ws.cell(row, 2).value.split()[0]} {ws.cell(row, 2).value.split()[1][0]}. {ws.cell(row, 2).value.split()[2][0]}.")
        print("АДРЕС1", ws.cell(row, 4).value)
        print("Телефон1", ws.cell(row, 5).value)
        print("Напряжение", str(ws.cell(row, 6).value).replace(".", ","))
        print("Мощность", str(ws.cell(row, 7).value).replace(".", ","))
        print("ПАСПОРТ1", ws.cell(row, 8).value)
        print("ТУ1", " ".join(ws.cell(row, 9).value.split()))
        print("НОМЕР1", ws.cell(row, 10).value)
        print("СУММА1", ws.cell(row, 12).value)
        print("Акт номер", ws.cell(row, 13).value)
        print("Тип", ws.cell(row, 14).value)
        print("Номинальный ток", ws.cell(row, 15).value)
        print("Заводской номер", ws.cell(row, 16).value)
        print("ДАТА1", f"{strftime('%d %m %Y', gmtime()).split()[0]} {months[strftime('%d %m %Y', gmtime()).split()[1]]} {strftime('%d %m %Y', gmtime()).split()[2]}")
        doc = DocxTemplate("1.docx")
        context = {'ФИО1': ws.cell(row, 3).value,
                   'ФИО3': ws.cell(row, 2).value,
                   "ФИО2": f"{ws.cell(row, 2).value.split()[0]} {ws.cell(row, 2).value.split()[1][0]}. {ws.cell(row, 2).value.split()[2][0]}.",
                   "АДРЕС1": ws.cell(row, 4).value,
                   "Телефон1": ws.cell(row, 5).value,
                   "Напряжение": str(ws.cell(row, 6).value).replace(".", ","),
                   "Мощность": str(ws.cell(row, 7).value).replace(".", ","),
                   "ПАСПОРТ1": ws.cell(row, 8).value,
                   "ТУ1": " ".join(ws.cell(row, 9).value.split()),
                   "НОМЕР1": ws.cell(row, 10).value,
                   "СУММА1": ws.cell(row, 12).value,
                   "Акт_номер": ws.cell(row, 13).value,
                   "Тип": ws.cell(row, 14).value,
                   "Номинальный_ток": ws.cell(row, 15).value,
                   "Заводской_номер": ws.cell(row, 16).value,
                   "ДАТА1": f"{strftime('%d %m %Y', gmtime()).split()[0]} {months[strftime('%d %m %Y', gmtime()).split()[1]]} {strftime('%d %m %Y', gmtime()).split()[2]}г."}
        doc.render(context)
        doc.save(f"{ws.cell(row, 2).value}.docx")



# doc = DocxTemplate("11.docx")
# context = {'ФИО1': "И.И.Иванов"}
# doc.render(context)
# doc.save("шаблон-final.docx")