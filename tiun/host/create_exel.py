
from docxtpl import DocxTemplate
import openpyxl
from time import gmtime, strftime
import docx2txt

months = {"01": "января", "02": "февраля", "03": "марта", "04": "апреля", "05": "мая", "06": "июня",
              "07": "июля", "08": "августа", "09": "сентября", "10": "октября", "11": "ноября", "12": "декабря"}

wookbook = openpyxl.load_workbook("Список заявок СК Репка.xlsx")
ws = wookbook.active

for row in range(2, ws.max_row):
    try:
        print("ФИО1",  ws.cell(row, 1).value, ws.cell(row, 2).value)

        text = docx2txt.process(f"{ws   .cell(row, 2).value}.docx")
        x = text.split()
        num = "".join(x[x.index("№")+1:x.index("от")])
        typ = "".join(x[x.index("Тип")+1:x.index("Заводской")])
        nom = "".join(x[x.index("А")+1:x.index("Номинальное")])
        zav = "".join(x[x.index("номер")+1:x.index("Класс")])

        print(num)
        print(typ)
        print(nom)
        print(zav)

        ws[f'M{ws.cell(row, 1).value+1}'].value = num
        ws[f'N{ws.cell(row, 1).value+1}'].value = typ
        ws[f'O{ws.cell(row, 1).value+1}'].value = nom
        ws[f'P{ws.cell(row, 1).value+1}'].value = zav
        wookbook.save('Список заявок СК Репка.xlsx')
    except:
        pass

